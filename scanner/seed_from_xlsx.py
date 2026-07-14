"""One-time seeder: convert the uploaded press-list xlsx into data/press_list.json."""
import json, re, sys
from datetime import date
import openpyxl

XLSX = sys.argv[1]
OUT = sys.argv[2]

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Master Press List"]

contacts = []
seen_emails = set()
rows = list(ws.iter_rows(values_only=True))
# header row is the one starting with 'Name'
start = next(i for i, r in enumerate(rows) if r[0] == "Name") + 1
for r in rows[start:]:
    name, outlet, email, category, status, beat, phone, deliv, relevance = (list(r) + [None]*9)[:9]
    if not email:
        continue
    email = str(email).strip().lower()
    if email in seen_emails:
        continue
    seen_emails.add(email)
    verified = isinstance(status, str) and "ACTIVE" in status
    contacts.append({
        "id": email,
        "name": (name or "").strip() or None,
        "outlet": (outlet or "").strip() or None,
        "email": email,
        "category": (category or "").strip() or None,
        "beat": (beat or "").strip() or None,
        "phone": str(phone).strip() if phone else None,
        "status": "active" if verified else "unverified",
        "status_note": (status or "").strip() or None,
        "deliverability_note": (deliv or "").strip() or None,
        "relevance": (relevance or "").strip() or None,
        "source": "seed_spreadsheet",
        "added": str(date.today()),
        "last_byline": None,          # date of most recent byline the scanner saw
        "last_byline_url": None,
        "byline_count": 0,
        "seeded": True,
    })

# Removed & flagged tab -> removed list (only the truly removed ones)
removed = []
ws2 = wb["Removed & flagged"]
rows2 = list(ws2.iter_rows(values_only=True))
start2 = next(i for i, r in enumerate(rows2) if r[0] == "Name") + 1
for r in rows2[start2:]:
    name, outlet, email, action = (list(r) + [None]*4)[:4]
    if not email or (action and str(action).upper().startswith("KEPT")):
        continue
    removed.append({
        "name": (name or "").strip() or None,
        "outlet": (outlet or "").strip() or None,
        "email": str(email).strip().lower(),
        "reason": (action or "").strip(),
        "removed_on": str(date.today()),
        "source": "seed_spreadsheet",
    })

out = {
    "generated": str(date.today()),
    "contacts": contacts,
    "removed": removed,
}
with open(OUT, "w") as f:
    json.dump(out, f, indent=1, ensure_ascii=False)
print(f"{len(contacts)} contacts, {len(removed)} removed")
